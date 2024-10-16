import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Criando uma nova planilha
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Checklist do Projeto"

# Dados do checklist (substitua conforme necessário)
checklist = [
    ["Fase 1: Configuração do Ambiente e Estudo das APIs", "Status"],
    ["1.1 Configuração do Ambiente de Desenvolvimento", "A Fazer"],
    ["Python: Instale o ambiente de desenvolvimento e configure ambiente virtual", "A Fazer"],
    ["Java: Configure o projeto e certifique-se que o Java 11+ está instalado", "A Fazer"],
    ["Frameworks: Configurar Flask/FastAPI e JavaFX/Swing", "A Fazer"],
    ["1.2 Estudo e Configuração das APIs do ML e Shopee", "A Fazer"],
    ["Mercado Livre: Estude a documentação e gere credenciais", "A Fazer"],
    ["Shopee: Estude a documentação e crie sandbox para testes", "A Fazer"],
    ["1.3 Testes Locais", "A Fazer"],
    ["Realize chamadas simples para as APIs", "A Fazer"],

    ["Fase 2: Implementação da API Python para Consumo de Dados", "Status"],
    ["2.1 Desenvolvimento do Módulo Python para Consumo das APIs", "A Fazer"],
    ["Funções para Mercado Livre: Recuperar pedidos, produtos e inventário", "A Fazer"],
    ["Funções para Shopee: Recuperar pedidos, produtos e inventário", "A Fazer"],
    ["2.2 Criação do Backend com Flask ou FastAPI", "A Fazer"],
    ["Implemente endpoints para pedidos, produtos e status de vendas", "A Fazer"],
    ["2.3 Testes Unitários e de Integração", "A Fazer"],
    ["Escreva testes para garantir que os dados são processados corretamente", "A Fazer"],

    ["Fase 3: Criação do Banco de Dados e Persistência", "Status"],
    ["3.1 Integração com Banco de Dados", "A Fazer"],
    ["Escolha e configure um banco de dados (MySQL, PostgreSQL, SQLite)", "A Fazer"],
    ["3.2 Salvando Dados Importantes", "A Fazer"],
    ["Defina os dados que serão persistidos e configure inserções no banco", "A Fazer"],
    ["3.3 Configuração de Atualizações Automáticas", "A Fazer"],
    ["Configure updates periódicos dos dados usando bibliotecas como schedule", "A Fazer"],

    ["Fase 4: Exibição dos Dados via API REST", "Status"],
    ["4.1 Expondo Dados com a API REST", "A Fazer"],
    ["Implemente endpoints REST para acessar dados processados", "A Fazer"],
    ["4.2 Testes de Performance", "A Fazer"],
    ["Teste performance da API em cenários simulados de carga", "A Fazer"],

    ["Fase 5: Implementação da UI Java e Integração", "Status"],
    ["5.1 Desenvolvimento da UI Java", "A Fazer"],
    ["Desenvolva a interface com JavaFX ou Swing", "A Fazer"],
    ["5.2 Consumindo a API Python na UI Java", "A Fazer"],
    ["Use HttpClient/Retrofit para consumir dados da API na UI", "A Fazer"],
    ["5.3 Tratamento de Erros e Manuseio de Respostas", "A Fazer"],
    ["Implemente tratamento de erros e ajuste a estrutura para performance", "A Fazer"],

    ["Fase 6: Testes Finais e Otimização", "Status"],
    ["6.1 Testes de Integração Completa", "A Fazer"],
    ["Teste end-to-end entre APIs, processamento e exibição em Java", "A Fazer"],
    ["6.2 Otimização do Desempenho", "A Fazer"],
    ["Identifique gargalos e otimize o sistema para alta performance", "A Fazer"],
    ["6.3 Implementação de Segurança", "A Fazer"],
    ["Adicione autenticação para proteger os endpoints da API", "A Fazer"],

    ["Fase 7: Implementação e Manutenção Contínua", "Status"],
    ["7.1 Deploy do Sistema", "A Fazer"],
    ["Faça o deploy do servidor Python em ambiente de produção", "A Fazer"],
    ["7.2 Monitoramento e Manutenção", "A Fazer"],
    ["Implemente logs e continue monitorando o sistema", "A Fazer"],
]

# Definir a linha inicial
row = 1

# Preenchendo os dados com célula mesclada para as atividades
for i in range(0, len(checklist)):
    # Verifica se é uma fase/título
    if "Fase" in checklist[i][0]:
        # Mescla as células da coluna 1 até a coluna 2 para a atividade principal
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1).value = checklist[i][0]
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    else:
        # Inserindo as subtarefas nas colunas correspondentes
        ws.cell(row=row, column=1).value = checklist[i][0]
        ws.cell(row=row, column=2).value = checklist[i][1]

    # Aumenta a linha para a próxima seção
    row += 1

# Ajustar a largura das colunas
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

# Salvando o arquivo Excel
wb.save("Checklist_Projeto_APIs_ML_Shopee_Completo.xlsx")

print("Arquivo Excel criado com sucesso!")
